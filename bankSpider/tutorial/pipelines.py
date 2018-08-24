# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import Workbook

class YinJianJuPipeline(object):
    wb = Workbook()
    ws = wb.active
    ws.append(['书文号', '个人姓名', '单位名称', '法定代表人', '违规事实', '处罚依据','处罚决定','机关名称','决定日期','发布时间','文章来源','文章类型','链接'])

    def process_item(self, item, spider):
        line = [item['wenshuhao'],item['geren'], item['mingcheng'], item['fadingren'],
                item['weiguishishi'],item['chufayiju'],item['chufajueding'],item['jiguanmingcheng'], item['chufariqi']
                ,item['pubDate'],item['source'], item['type'],item['link']]
        self.ws.append(line)  # 将数据以行的形式添加到xlsx中
        self.wb.save('yinjianju.xlsx')  # 保存xlsx文件
        return item

class YinJianHuiJiGuanPipeline(object):
    wb = Workbook()
    ws = wb.active
    ws.append(['书文号', '个人姓名', '单位名称', '法定代表人', '违规事实', '处罚依据','处罚决定','机关名称','决定日期','发布时间', '文章来源','文章类型','链接'])

    def process_item(self, item, spider):
        line = [item['wenshuhao'],item['geren'], item['mingcheng'], item['fadingren'],
                item['weiguishishi'],item['chufayiju'],item['chufajueding'],item['jiguanmingcheng'], item['chufariqi']
                ,item['pubDate'],item['source'], item['type'],item['link']]
        self.ws.append(line)  # 将数据以行的形式添加到xlsx中
        self.wb.save('yinjianhuijiguan.xlsx')  # 保存xlsx文件
        return item

class YinJianFenJuPipeline(object):
    wb = Workbook()
    ws = wb.active
    ws.append(['书文号', '个人姓名', '单位名称', '法定代表人', '违规事实', '处罚依据','处罚决定','机关名称','决定日期','发布时间','文章来源','文章类型','链接'])

    def process_item(self, item, spider):
        line = [item['wenshuhao'],item['geren'], item['mingcheng'], item['fadingren'],
                item['weiguishishi'],item['chufayiju'],item['chufajueding'],item['jiguanmingcheng'], item['chufariqi']
                ,item['pubDate'],item['source'], item['type'],item['link']]
        self.ws.append(line)  # 将数据以行的形式添加到xlsx中
        self.wb.save('yinjianfenju.xlsx')  # 保存xlsx文件
        return item
